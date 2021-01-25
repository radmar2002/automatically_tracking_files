import time
import os
import sys
import sqlite3
import hashlib
import openpyxl
import socket

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlite3 import Error
from datetime import datetime


def getbasefile():
    """
    Returns the name of the SQLite DB file
    """
    return os.path.splitext(os.path.basename(__file__))[0]


def connectdb():
    """
    Connects to the SQLite DB
    """
    try:
        dbfile = getbasefile() + '.db'
        conn = sqlite3.connect(dbfile, timeout=2)
        print("Connection is established: Database is created on disk")
        return conn
    except Error as e:
        print('Connection went wrong:', e)


def corecursor(conn, query, args=None):
    """
    Run a SQLite DB cursor
    """
    cursor = None
    args = args or []
    try:
        cursor = conn.execute(query, args)
        return cursor
    except Error as e:
        print(e)
    return cursor


def tableexists(table):
    """
    Checks if a SQLite DB Table exists
    """
    result = False
    try:
        conn = connectdb()
        if conn != None:
            query = "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?"
            args = (table,)
            cursor = corecursor(conn, query, args)
            rows = cursor.fetchall()
            numrows = len(list(rows))
            if numrows > 0:
                result = True
        print('Table exist:', result)
    except Error as e:
        print(e)
    finally:
        cursor.close()
        if conn != None:
            conn.commit()
            conn.close()
            print("Closed connection to the database successfully")
    return result


def createhashtable():
    """
    Creates a SQLite DB Table
    Function that can create a file-level tracking database table
    on a local SQLite instance.
    """
    result = False
    query = "CREATE TABLE files (id INTEGER PRIMARY KEY, fname VARCHAR(255) NOT NULL, md5 BLOB NOT NULL )"
    try:
        conn = connectdb()
        if conn != None:
            if tableexists('files') == False:
                try:
                    cursor = corecursor(conn, query)
                    cursor.close()
                    result = True
                    print('Created a SQLite DB Table!')
                except Error as e:
                    print('Create a SQLite DB Table went wrong: ', e)
    except Error as e:
        print(e)
    finally:
        if conn != None:
            conn.commit()
            conn.close()
            print("Closed connection to the database successfully")
    return result


def createhashtableidx():
    """
    Creates a SQLite DB Table Index
    Function that create a file-level tracking table index
    on a local SQLite instance
    """
    result = False
    table = 'files'
    query = 'CREATE INDEX idxfile ON files (fname)'  # Finish this query ...
    try:
        conn = connectdb()
        if conn != None:
            if tableexists(table) == False:
                createhashtable()
            else:
                try:
                    cursor = corecursor(conn, query)
                    # cursor.close()
                    result = True
                    print('Create a SQLite DB Table INDEX!')
                except Error as e:
                    print('Create a SQLite DB Table INDEX went wrong: ', e)
    except Error as e:
        print(e)
    finally:
        if conn != None:
            conn.commit()
            conn.close()
            print("Closed connection to the database successfully")
    return result


def runcmd(query, args=None):
    """
    Run a specific command on the SQLite DB
    """
    args = args or []
    result = False
    try:
        conn = connectdb()
        if conn != None:
            if tableexists('files'):
                try:
                    cursor = conn.cursor()
                    cursor.execute(query, args)
                    conn.commit()
                    cursor.close()
                    result = True
                except Exception as e:
                    print("Query execution error: ", e)
    except Exception as e:
        print(e)
    return result


def updatehashtable(fname, md5):
    """
    Update the SQLite File Table
    """
    query = "UPDATE files SET md5 = ? WHERE fname = ?"
    runcmd(query, (md5, fname))


def inserthashtable(fname, md5):
    """
    Insert into the SQLite File Table
    """
    query = "INSERT INTO files (fname, md5) VALUES (?,?)"
    runcmd(query, (fname, md5))


def setuphashtable(fname, md5):
    """
    Setup's the Hash Table
    """
    createhashtable()
    createhashtableidx()
    inserthashtable(fname, md5)


def md5indb(fname):
    """
    Checks if md5 hash tag exists in the SQLite DB
    """
    query = "SELECT md5 FROM files WHERE fname = ?"
    try:
        conn = connectdb()
        if not conn is None:
            if tableexists('files'):
                try:
                    cursor = conn.cursor()
                    args = (fname, )
                    md5row = corecursor(conn, query, args).fetchone()
                    if md5row:
                        return md5row[0]
                except Error as e:
                    print(e)
                finally:
                    cursor.close()
                    if conn != None:
                        conn.commit()
                        conn.close()
                        print("Closed connection to the database successfully")
    except Error as e:
        print(e)
    return None


def haschanged(fname, md5):
    """
    Checks if a file has changed
    """
    result = None
    fileMD5inDB = md5indb(fname)
    if fileMD5inDB is None:
        setuphashtable(fname, md5)
        result = 'IS_SETUP'
        return result
    elif fileMD5inDB != md5:
        updatehashtable(fname, md5)
        result = 'CHANGED'
        return result
    elif fileMD5inDB == md5:
        updatehashtable(fname, md5)
        result = 'NOT_CHANGED'
        return result
    else:
        raise ValueError('A MD5 corner case might happened.')
    return result


def getfileext(fname):
    """
    Get the file name extension
    """
    return '.' + os.path.splitext(fname)[1][1:]


def getmoddate(fname):
    """
    Get file modified date
    """
    try:
        mtime = os.path.getmtime(fname)
    except Error as e:
        print(e)
    return mtime


def md5short(fname):
    """
    Get md5 file hash tag
    """
    with open(fname, 'r') as open_file:
        content = open_file.read()
        hasher = hashlib.md5(content.encode('utf-8'))
    md5value = hasher.hexdigest()
    return md5value


def loadflds():
    """
    Write a Python function that can load and parse the configuration file.
    This function should return the list of folders and list of extensions
    to exclude for each folder (where applicable).
    """
    flds = []
    ext = []
    #config = getbasefile() + '.ini'
    config = os.path.join(os.getcwd(), 'filechanges.ini')
    if os.path.isfile(config):
        cfile = open(config, 'r')
        # Parse each config file line and get the folder and extensions
        for dirline in cfile:
            if len(dirline.split("|")) == 2:
                extensions = dirline.split("|")[1]
                entensions = list(set(extensions.split(",")))
                entensions = [e.replace('\n', '') for e in entensions]
                ext.append(entensions)
                flds.append(dirline.split("|")[0])
            else:
                flds.append(dirline)
                ext.append([])
    return flds, ext


def checkfilechanges(folder, exclude, ws=None):
    changed = False
    """Checks for files changes"""
    for subdir, dirs, files in os.walk(folder, topdown=True):
        for fname in files:
            origin = os.path.normpath(os.path.join(subdir, fname))
            if os.path.isfile(origin):
                # Get file extension and check if it is not excluded
                fileext = getfileext(origin)
                if fileext not in exclude:
                    print('===>', origin)
                    # Get the file’s md5 hash
                    filemd5 = md5short(origin)
                    print('File’s md5 hash is:', filemd5, md5indb(origin))
                    # If the file has changed, add it to the Excel report
                    file_changed = haschanged(origin, filemd5)
                    if file_changed != 'NOT_CHANGED':
                        changed = True
                        with open('REPORT_FILE.csv', "a") as file_object:
                            # Append change log at the end of file
                            file_object.write(
                                file_changed+", " +
                                str(getmoddate(origin)) + ', ' + origin + "\n"
                            )
                    print('file has changed', haschanged(origin, filemd5))
    return changed


def runfilechanges(ws=None):
    changed = False
    # Invoke the function that loads and parses the config file
    currentpaths, bannedextensions = loadflds()
    for i, fld in enumerate(currentpaths):
        print('List banned extensions: ', bannedextensions[i], '<--->', fld)
        # Invoke the function that checks each folder for file changes
        if checkfilechanges(fld, bannedextensions[i], ws):
            changed = True
    return changed


def getdt(frmt):
    today = datetime.today()
    return today.strftime(frmt)


def startxlsreport(cnf):
    # Create the workbook, get the hostname and current DateTime
    xls = getbasefile() + '.xlsx'
    if os.path.exists(xls):
        wb = Workbook() if cnf == True else openpyxl.load_workbook(xls)
    else:
        wb = Workbook()
    ws = wb.active
    ws.title = socket.gethostname()
    st = getdt("%d-%b-%Y %H_%M_%S") if cnf == True else ''
    return wb, ws, st


def endxlsreport(wb, st):
    if st != '':
        dt = ' from ' + st + ' to ' + getdt("%d-%b-%Y %H_%M_%S")
        fn = getbasefile() + dt + '.xlsx'
    else:
        fn = getbasefile() + '.xlsx'
    wb.save(fn)


def headerxlsreport(ws):
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Full File Name")
    ws.cell(row=1, column=3, value="Folder Name")
    ws.cell(row=1, column=4, value="Date")
    ws.cell(row=1, column=5, value="Time")

    ft = Font(color="000000", bold=True)
    ws["A1"].font = ft
    ws["A2"].font = ft
    ws["A3"].font = ft
    ws["A4"].font = ft
    ws["A6"].font = ft


def getlastrow(ws):
    rw = 1
    for cell in ws["A"]:
        if cell.value is None:
            break
        else:
            rw += 1
    return rw


def rowxlsreport(ws, fn, ffn, fld, d, t):
    row = getlastrow(ws)
    ws.cell(row=row, column=1, value=fn)
    ws.cell(row=row, column=2, value=ffn)
    ws.cell(row=row, column=3, value=fld)
    ws.cell(row=row, column=4, value=d)
    ws.cell(row=row, column=5, value=t)


def execute(args):
    changed = False
    # Start the creation of the Excel report
    if len(args) > 1:
        if args[1].lower() == '--loop':
            if len(args) == 3:
                cnf = True if args[2].lower() == '--cnf' else False
            else:
                cnf = False
            wb, ws, st = startxlsreport(cnf)
            try:
                while True:
                    changed = runfilechanges(ws)
            except KeyboardInterrupt:
                print('Program stopped!!')
                if changed:
                    endxlsreport(wb, st)
    else:
        wb, ws, st = startxlsreport()
        changed = runfilechanges(ws)
        if changed:
            endxlsreport(wb, st)


"""Check functionality"""
if __name__ == "__main__":
    print(sys.argv)
    execute(sys.argv)
